import csv
import os.path
import re
import sys
import time
import openpyxl
# use the yaml library
import yaml


def junk():
    return 1


# data-oriented programming
# - figure out your inputs and outputs
# - try to make your functions reusable (for yourself and other people)
#    - no sys.exit()
#    - try not to many unnecessary changes that might prevent re-usability
# - don't mix input/output with your processing (hurts testing, and makes it hard to re-use)
def verify_num_args_passed(args):
    # Ensure 2 arguments are passed into program
    if args is None or len(args) < 3:
        print('\n***ERROR*** 2 arguments are required.\n')
        print('required: Argument 1: the path to the input CSV file itself')
        print('required: Argument 2: the path to the YAML output file\n')
        print('Quietly terminating myself.....')
        return False
    else:
        return True


def verify_paths_valid(args):
    # Verify the paths that were passed are valid:
    # (arg[1] is path/filename to csv input; arg[2] is path/filename to output yml file
    if not (os.path.isfile(args[1])):
        print(sys.argv[1], ' filename does not exist.\nQuietly terminating myself...')
        return False
    else:
        return True


def verify_out_path(path):
    if not (os.path.isdir(path)):
        print(out_path, ' is not a valid path.\nQuietly terminating myself...')
        sys.exit()


def verify_fname_valid_chars(ofile):
    # Verify filename to write to contain valid filename chars.
    # NOT perfect as system chaining operators will override the input.
    # You get the idea of where my head is at.
    match = re.search(r'[#<$+%>!`&*|{?=}/:r"\"@^]', ofile)

    if match is not None:
        print(match.group(), 'is not a valid character for a filename.\nQuietly terminating myself...')
        sys.exit()


# - two inputs
#   - input CSV path
#   - output YAML path
#   - delimiter


# DOP: this should actually return a list of dictionaries
#      a separate function should dump the list of dicts to a file

# side effects are stuff that change the "real world", like files or databases
# prefer many functions that are "pure" (i.e., don't change the real world), with a
# few functions that do
def process_csv(delimiter=','):
    with open(sys.argv[1], mode='r') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter)
        header = next(csv_reader)
        print(f'Total column headers:    {len(header)}')  # total sols

        # Read data contents of csv file into memory
        each_rows = []
        start_time = time.time()

        for row in csv_reader:
            each_rows.append(row)
        print(f'Total record entries:    {len(each_rows)}\n')

        with open(sys.argv[2], mode='w') as yaml_out_file:
            yaml_out_file.write('---\n')

            for idx, row_data in enumerate(each_rows):
                if len(row_data) != len(header):
                    print(
                        f'******** Potential error or data missing in row {idx + 1}. Only {len(row_data)} '
                        f'values found - should be {len(header)} values. *********\nContinuing to process data')
                    print(f'processing {row_data}')

                    for idx2 in range(len(row_data)):
                        if idx2 == 0:
                            yaml_out_file.write(f'- {row_data[idx2]}:\n')
                        else:
                            yaml_out_file.write(f'    {header[idx2]}: {row_data[idx2]}\n')
                else:
                    print(f'processing {row_data}')
                    for idx2 in range(len(header)):
                        if idx2 == 0:
                            yaml_out_file.write(f'- {row_data[idx2]}:\n')
                        else:
                            yaml_out_file.write(f'    {header[idx2]}: {row_data[idx2]}\n')

            yaml_out_file.write('...')

        end_time = time.time()
        time_elapsed = (end_time - start_time)

        print(f'\nData processing completed in {round(time_elapsed * 1000, 3)} ms.\n')


def process_xlsx():
    wb = openpyxl.load_workbook(sys.argv[1])

    # No specification given for which sheet to use, so defaulting to first sheet in Excel file
    sheet = wb.worksheets[0]
    row_count = sheet.max_row - 1
    col_count = sheet.max_column - 1

    print(f'Total column headers:    {col_count}\nTotal record entries:    {row_count}\n')

    header = []
    start_time = time.time()

    # Get headers
    for row_header in sheet.iter_rows(max_row=1):
        for cell in row_header:
            if cell.value is None:  # for whatever reason, openpyxl iterates max + 1 when iter_rows() is run
                continue
            else:
                header.append(cell.value)

    # Process row data
    with open(sys.argv[2], mode='w') as yaml_out_file:
        yaml_out_file.write('---\n')

        for row_values in sheet.iter_rows(max_col=col_count, min_row=2):
            for idx, header_val in enumerate(header):
                if idx == 0:
                    data = f'- {row_values[idx].value}:\n'
                    yaml_out_file.write(data)
                else:
                    print(header[idx], row_values[idx].value)
                    data = f'    {header[idx]}: {row_values[idx].value}\n'
                    yaml_out_file.write(data)

        yaml_out_file.write('...')

    end_time = time.time()
    time_elapsed = (end_time - start_time)

    print(f'\nData processing completed in {round(time_elapsed * 1000, 3)} ms.\n')


if __name__ == "__main__":
    print('\n\n======  CSV/XLSX to YAML converter   ====== ')

    result_num_args_passed = verify_num_args_passed(sys.argv)

    if result_num_args_passed is False:
        sys.exit()

    result_paths_valid = verify_paths_valid(sys.argv)

    if result_paths_valid is False:
        sys.exit()

    out_path, out_file = os.path.split(sys.argv[2])  # head returns path, tail returns filename
    yaml_out_path = os.path.isdir(out_path)

    verify_out_path(out_path)
    verify_fname_valid_chars(out_file)

    print(f'Input file given:        {sys.argv[1]}  - VALID INPUT')
    print(f'Output file path valid:  {out_file}     - VALID INPUT & will be written to: {out_path}')

    input_type = os.path.splitext(sys.argv[1])[1]

    if input_type == '.csv':
        process_csv()
    elif input_type == '.xlsx':
        process_xlsx()
    else:
        print(f'{input_type} is not a valid extension and the file {sys.argv[1]} cannot be processed.')
        print('Quietly terminating myself...')
        sys.exit()
